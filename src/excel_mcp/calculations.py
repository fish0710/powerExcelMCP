from typing import Any, Dict, List, Union
import logging
from pathlib import Path
import openpyxl

from .workbook import get_or_create_workbook
from .cell_utils import validate_cell_reference, parse_cell_range
from .exceptions import ValidationError, CalculationError
from .validation import validate_formula

logger = logging.getLogger(__name__)

def apply_formula(
    filepath: Path | str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> Dict[str, Any]:
    """Apply formula to cell in Excel sheet"""
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(filepath)
        
        # Check if sheet exists
        if sheet_name not in workbook.sheetnames:
            raise CalculationError(f"Sheet '{sheet_name}' not found")
            
        worksheet = workbook[sheet_name]
        
        # Parse cell coordinates
        try:
            cell_coords = parse_cell_range(f"{cell}:{cell}")
            if not cell_coords or not all(coord is not None for coord in cell_coords[:2]):
                raise CalculationError(f"Invalid cell reference: {cell}")
            row, col = cell_coords[0], cell_coords[1]
        except ValueError as e:
            raise CalculationError(f"Invalid cell format: {str(e)}")
            
        # Apply formula
        target_cell = worksheet.cell(row=row, column=col)
        
        # Remove equal sign if present
        if formula.startswith('='):
            formula = formula[1:]
            
        target_cell.value = f"={formula}"
        
        # Save workbook
        workbook.save(filepath)
        
        return {"message": f"Applied formula to {cell}"}
    except (CalculationError, ValidationError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to apply formula: {e}")
        raise CalculationError(str(e))

def batch_apply_formulas(
    filepath: Path | str,
    sheet_name: str,
    formulas: List[Dict[str, str]]
) -> Dict[str, Any]:
    """批量应用多个公式到不同单元格，提高性能
    
    Args:
        filepath: Excel文件路径
        sheet_name: 工作表名称
        formulas: 公式列表，每个元素是包含cell和formula键的字典
        
    Returns:
        包含应用结果的字典
    """
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(filepath)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise CalculationError(f"工作表 '{sheet_name}' 不存在")
            
        worksheet = workbook[sheet_name]
        
        # 追踪成功和失败的公式
        applied_count = 0
        failed_count = 0
        failures = []
        
        # 遍历并应用所有公式
        for idx, formula_obj in enumerate(formulas):
            if not isinstance(formula_obj, dict):
                failed_count += 1
                failures.append(f"索引 {idx}: 公式对象不是字典")
                continue
                
            cell = formula_obj.get('cell')
            formula = formula_obj.get('formula')
            
            if not cell or not formula:
                failed_count += 1
                failures.append(f"索引 {idx}: 缺少cell或formula字段")
                continue
            
            try:
                # 解析单元格坐标
                cell_coords = parse_cell_range(f"{cell}:{cell}")
                if not cell_coords or not all(coord is not None for coord in cell_coords[:2]):
                    failed_count += 1
                    failures.append(f"索引 {idx}: 无效的单元格引用: {cell}")
                    continue
                    
                row, col = cell_coords[0], cell_coords[1]
                
                # 应用公式
                target_cell = worksheet.cell(row=row, column=col)
                
                # 如果公式以等号开头则移除
                if isinstance(formula, str) and formula.startswith('='):
                    formula = formula[1:]
                    
                target_cell.value = f"={formula}"
                applied_count += 1
                
            except Exception as e:
                failed_count += 1
                failures.append(f"索引 {idx}: {str(e)}")
        
        # 保存工作簿
        workbook.save(filepath)
        
        result = {
            "message": f"批量应用了 {applied_count} 个公式，{failed_count} 个失败",
            "applied_count": applied_count,
            "failed_count": failed_count
        }
        
        if failures:
            result["failures"] = failures[:10]  # 只返回前10个失败，避免返回过多数据
            if len(failures) > 10:
                result["failures"].append(f"... 还有 {len(failures) - 10} 个失败未显示")
                
        return result
        
    except (CalculationError, ValidationError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"批量应用公式失败: {e}")
        raise CalculationError(str(e))

def apply_array_formula(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    formula: str
) -> Dict[str, Any]:
    """将数组公式应用到单元格区域
    
    Args:
        filepath: Excel文件路径
        sheet_name: 工作表名称
        start_cell: 开始单元格
        end_cell: 结束单元格
        formula: 数组公式
        
    Returns:
        包含应用结果的字典
    """
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(filepath)
        
        # 检查工作表是否存在
        if sheet_name not in workbook.sheetnames:
            raise CalculationError(f"工作表 '{sheet_name}' 不存在")
            
        worksheet = workbook[sheet_name]
        
        # 解析单元格范围
        try:
            cell_range = parse_cell_range(f"{start_cell}:{end_cell}")
            if not cell_range or len(cell_range) != 4 or any(coord is None for coord in cell_range):
                raise CalculationError(f"无效的单元格范围: {start_cell}:{end_cell}")
                
            start_row, start_col, end_row, end_col = cell_range
        except ValueError as e:
            raise CalculationError(f"无效的单元格格式: {str(e)}")
        
        # 统计要应用的单元格数量
        cell_count = (end_row - start_row + 1) * (end_col - start_col + 1)
        
        # 如果单元格数量过多，可能会影响性能
        if cell_count > 1000:
            logger.warning(f"将数组公式应用到大范围 ({cell_count} 个单元格)，这可能会很耗时")
            
        # 如果公式以等号开头则移除
        if formula.startswith('='):
            formula = formula[1:]
            
        # 将数组公式应用到每个单元格
        # 注意：真正的数组公式在Excel中是一次性应用到整个范围的，
        # 但openpyxl不完全支持这个功能，所以我们需要循环应用到每个单元格
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = f"={formula}"
                
        # 保存工作簿
        workbook.save(filepath)
        
        return {
            "message": f"已将数组公式应用到区域 {start_cell}:{end_cell} ({cell_count} 个单元格)"
        }
        
    except (CalculationError, ValidationError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"应用数组公式失败: {e}")
        raise CalculationError(str(e))

def calculate_with_pandas(
    filepath: Path | str,
    sheet_name: str,
    range_ref: str,
    operation: str,
    output_cell: str = None
) -> Dict[str, Any]:
    """使用pandas进行高性能计算
    
    Args:
        filepath: Excel文件路径
        sheet_name: 工作表名称
        range_ref: 数据范围引用，如"A1:C10"
        operation: 执行的操作，如"sum", "mean", "max", "min", "count", "corr"
        output_cell: 将结果输出到的单元格，可选
        
    Returns:
        包含计算结果的字典
    """
    try:
        import pandas as pd
        import numpy as np
        
        # 解析范围
        try:
            if ':' not in range_ref:
                raise CalculationError(f"无效的范围引用: {range_ref}，必须是类似'A1:C10'的格式")
                
            start_cell, end_cell = range_ref.split(':')
            range_coords = parse_cell_range(f"{start_cell}:{end_cell}")
            
            if not range_coords or len(range_coords) != 4 or any(coord is None for coord in range_coords):
                raise CalculationError(f"无效的单元格范围: {range_ref}")
                
            start_row, start_col, end_row, end_col = range_coords
        except ValueError as e:
            raise CalculationError(f"无效的单元格格式: {str(e)}")
            
        # 从Excel读取数据到DataFrame
        df = pd.read_excel(
            filepath, 
            sheet_name=sheet_name,
            header=None  # 不使用标题行，所有数据都会被读取为数值
        )
        
        # 调整为指定范围（DataFrame是0索引，而Excel是1索引）
        df = df.iloc[start_row-1:end_row, start_col-1:end_col]
        
        # 根据操作类型执行计算
        result = None
        
        # 统计操作（应用于整个数据集）
        if operation == "sum":
            result = df.values.sum()
        elif operation == "mean":
            result = df.values.mean()
        elif operation == "min":
            result = df.values.min()
        elif operation == "max":
            result = df.values.max()
        elif operation == "count":
            result = df.count().sum()  # 统计非NaN值的数量
        elif operation == "median":
            result = np.median(df.values)
        elif operation == "std":
            result = df.values.std()
        elif operation == "var":
            result = df.values.var()
            
        # 相关性分析
        elif operation == "corr":
            # 检查是否有足够的列进行相关性分析
            if df.shape[1] < 2:
                raise CalculationError("相关性分析至少需要2列数据")
                
            # 计算相关性矩阵
            corr_matrix = df.corr().to_dict()
            result = corr_matrix
            
        # 不支持的操作
        else:
            raise CalculationError(f"不支持的操作: {operation}。支持的操作: sum, mean, min, max, count, median, std, var, corr")
            
        # 如果指定了输出单元格，写入结果
        if output_cell and not isinstance(result, dict):  # 字典结果（如相关性矩阵）不能写入单元格
            workbook = openpyxl.load_workbook(filepath)
            
            if sheet_name not in workbook.sheetnames:
                raise CalculationError(f"工作表 '{sheet_name}' 不存在")
                
            worksheet = workbook[sheet_name]
            
            # 解析输出单元格
            try:
                output_coords = parse_cell_range(f"{output_cell}:{output_cell}")
                if not output_coords or not all(coord is not None for coord in output_coords[:2]):
                    raise CalculationError(f"无效的输出单元格引用: {output_cell}")
                    
                out_row, out_col = output_coords[0], output_coords[1]
            except ValueError as e:
                raise CalculationError(f"无效的输出单元格格式: {str(e)}")
                
            # 写入结果
            output = worksheet.cell(row=out_row, column=out_col)
            output.value = result
            
            # 保存工作簿
            workbook.save(filepath)
            
        return {
            "message": f"使用pandas在范围 {range_ref} 上计算 {operation} 完成",
            "result": result
        }
        
    except ImportError:
        raise CalculationError("需要安装pandas和numpy库: pip install pandas numpy")
    except (CalculationError, ValidationError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"使用pandas计算失败: {e}")
        raise CalculationError(str(e))