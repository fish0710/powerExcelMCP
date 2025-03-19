from pathlib import Path
from typing import Any
import logging

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .exceptions import DataError
from .cell_utils import parse_cell_range

logger = logging.getLogger(__name__)

def read_excel_range(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    preview_only: bool = False
) -> list[dict[str, Any]]:
    """Read data from Excel range with optional preview mode"""
    try:
        wb = load_workbook(filepath, read_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")
            
        ws = wb[sheet_name]

        # Parse start cell
        if ':' in start_cell:
            start_cell, end_cell = start_cell.split(':')
            
        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # For single cell, use same coordinates
            end_row, end_col = start_row, start_col

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            raise DataError(
                f"Start cell out of bounds. Sheet dimensions are "
                f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            )

        data = []
        # If it's a single cell or single row, just read the values directly
        if start_row == end_row:
            row_data = {}
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=start_row, column=col)
                col_name = f"Column_{col}"
                row_data[col_name] = cell.value
            if any(v is not None for v in row_data.values()):
                data.append(row_data)
        else:
            # Multiple rows - use header row
            headers = []
            for col in range(start_col, end_col + 1):
                cell_value = ws.cell(row=start_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")

            # Get data rows
            max_rows = min(start_row + 5, end_row) if preview_only else end_row
            for row in range(start_row + 1, max_rows + 1):
                row_data = {}
                for col, header in enumerate(headers, start=start_col):
                    cell = ws.cell(row=row, column=col)
                    row_data[header] = cell.value
                if any(v is not None for v in row_data.values()):
                    data.append(row_data)

        wb.close()
        return data
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range: {e}")
        raise DataError(str(e))

def write_data(
    filepath: str,
    sheet_name: str | None,
    data: list[dict[str, Any]] | None,
    start_cell: str = "A1",
    write_headers: bool = True,
) -> dict[str, str]:
    """Write data to Excel sheet with workbook handling"""
    try:
        if not data:
            raise DataError("No data provided to write")
            
        wb = load_workbook(filepath)

        # If no sheet specified, use active sheet
        if not sheet_name:
            sheet_name = wb.active.title
        elif sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # Validate start cell
        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        if len(data) > 0:
            # Check if first row of data contains headers
            first_row = data[0]
            has_headers = all(
                isinstance(value, str) and value.strip() == key.strip()
                for key, value in first_row.items()
            )
            
            # If first row contains headers, skip it when write_headers is True
            if has_headers and write_headers:
                data = data[1:]

            _write_data_to_worksheet(ws, data, start_cell, write_headers)

        wb.save(filepath)
        wb.close()

        return {"message": f"Data written to {sheet_name}", "active_sheet": sheet_name}
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write data: {e}")
        raise DataError(str(e))

def _write_data_to_worksheet(
    worksheet: Worksheet, 
    data: list[dict[str, Any]], 
    start_cell: str = "A1",
    write_headers: bool = True,
) -> None:
    """Write data to worksheet - internal helper function"""
    try:
        if not data:
            raise DataError("No data provided to write")

        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(x is not None for x in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Validate data structure
        if not all(isinstance(row, dict) for row in data):
            raise DataError("All data rows must be dictionaries")

        # Write headers if requested
        headers = list(data[0].keys())
        if write_headers:
            for i, header in enumerate(headers):
                cell = worksheet.cell(row=start_row, column=start_col + i)
                cell.value = header
                cell.font = Font(bold=True)
            start_row += 1  # Move start row down if headers were written

        # Write data
        for i, row_dict in enumerate(data):
            if not all(h in row_dict for h in headers):
                raise DataError(f"Row {i+1} is missing required headers")
            for j, header in enumerate(headers):
                cell = worksheet.cell(row=start_row + i, column=start_col + j)
                cell.value = row_dict.get(header, "")
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write worksheet data: {e}")
        raise DataError(str(e))

def batch_process_data(
    filepath: str, 
    sheet_name: str, 
    operations: list[dict], 
    save_intervals: int = 100
) -> dict:
    """使用pandas处理Excel数据，支持高性能的批量操作。
    
    Args:
        filepath: Excel文件路径
        sheet_name: 工作表名称
        operations: 操作列表，每个操作是一个字典，包含'type'和其他相关参数
                  支持的操作类型: 'filter', 'transform', 'aggregate', 'sort'
        save_intervals: 每处理多少操作保存一次，提高大数据处理的可靠性
        
    Returns:
        包含处理结果信息的字典
    """
    try:
        import pandas as pd
        
        # 读取Excel数据到pandas DataFrame
        df = pd.read_excel(filepath, sheet_name=sheet_name)
        original_row_count = len(df)
        
        # 执行操作
        for i, op in enumerate(operations):
            op_type = op.get('type')
            
            if op_type == 'filter':
                # 过滤操作示例: {'type': 'filter', 'column': 'Age', 'operator': '>', 'value': 30}
                column = op.get('column')
                operator = op.get('operator')
                value = op.get('value')
                
                if not all([column, operator, value is not None]):
                    raise DataError(f"过滤操作缺少必要参数: {op}")
                
                if column not in df.columns:
                    raise DataError(f"列 '{column}' 不存在")
                
                # 根据操作符应用过滤条件
                if operator == '==':
                    df = df[df[column] == value]
                elif operator == '!=':
                    df = df[df[column] != value]
                elif operator == '>':
                    df = df[df[column] > value]
                elif operator == '<':
                    df = df[df[column] < value]
                elif operator == '>=':
                    df = df[df[column] >= value]
                elif operator == '<=':
                    df = df[df[column] <= value]
                elif operator == 'contains':
                    df = df[df[column].astype(str).str.contains(str(value), na=False)]
                elif operator == 'startswith':
                    df = df[df[column].astype(str).str.startswith(str(value), na=False)]
                elif operator == 'endswith':
                    df = df[df[column].astype(str).str.endswith(str(value), na=False)]
                else:
                    raise DataError(f"不支持的过滤操作符: {operator}")
            
            elif op_type == 'transform':
                # 转换操作示例: {'type': 'transform', 'column': 'Salary', 'method': 'multiply', 'value': 1.1}
                column = op.get('column')
                method = op.get('method')
                value = op.get('value')
                new_column = op.get('new_column', column)  # 可选参数，默认覆盖原列
                
                if not all([column, method]):
                    raise DataError(f"转换操作缺少必要参数: {op}")
                
                if column not in df.columns:
                    raise DataError(f"列 '{column}' 不存在")
                
                # 应用转换
                if method == 'add':
                    df[new_column] = df[column] + value
                elif method == 'subtract':
                    df[new_column] = df[column] - value
                elif method == 'multiply':
                    df[new_column] = df[column] * value
                elif method == 'divide':
                    df[new_column] = df[column] / value
                elif method == 'round':
                    decimals = int(value) if value is not None else 0
                    df[new_column] = df[column].round(decimals)
                elif method == 'upper':
                    df[new_column] = df[column].astype(str).str.upper()
                elif method == 'lower':
                    df[new_column] = df[column].astype(str).str.lower()
                elif method == 'replace':
                    old_val = op.get('old_value', '')
                    new_val = op.get('new_value', '')
                    df[new_column] = df[column].astype(str).str.replace(str(old_val), str(new_val))
                else:
                    raise DataError(f"不支持的转换方法: {method}")
            
            elif op_type == 'sort':
                # 排序操作示例: {'type': 'sort', 'columns': ['Age', 'Salary'], 'ascending': [True, False]}
                columns = op.get('columns', [])
                ascending = op.get('ascending', True)
                
                if not columns:
                    raise DataError(f"排序操作缺少必要参数: {op}")
                
                for col in columns:
                    if col not in df.columns:
                        raise DataError(f"列 '{col}' 不存在")
                
                # 应用排序
                df = df.sort_values(by=columns, ascending=ascending)
            
            elif op_type == 'aggregate':
                # 聚合操作示例: {'type': 'aggregate', 'group_by': 'Department', 'agg_column': 'Salary', 'method': 'mean'}
                group_by = op.get('group_by', [])
                agg_column = op.get('agg_column')
                method = op.get('method', 'mean')
                
                if not isinstance(group_by, list):
                    group_by = [group_by]
                
                if not all([group_by, agg_column, method]):
                    raise DataError(f"聚合操作缺少必要参数: {op}")
                
                for col in group_by + [agg_column]:
                    if col not in df.columns:
                        raise DataError(f"列 '{col}' 不存在")
                
                # 应用聚合
                if method == 'mean':
                    df = df.groupby(group_by)[agg_column].mean().reset_index()
                elif method == 'sum':
                    df = df.groupby(group_by)[agg_column].sum().reset_index()
                elif method == 'count':
                    df = df.groupby(group_by)[agg_column].count().reset_index()
                elif method == 'min':
                    df = df.groupby(group_by)[agg_column].min().reset_index()
                elif method == 'max':
                    df = df.groupby(group_by)[agg_column].max().reset_index()
                else:
                    raise DataError(f"不支持的聚合方法: {method}")
            
            else:
                raise DataError(f"不支持的操作类型: {op_type}")
            
            # 定期保存，防止处理大量数据时内存溢出
            if (i + 1) % save_intervals == 0:
                temp_output = f"{filepath}.temp.xlsx"
                df.to_excel(temp_output, sheet_name=sheet_name, index=False)
                logger.info(f"临时保存进度，已处理 {i + 1}/{len(operations)} 个操作")
        
        # 将处理后的DataFrame写回Excel
        with pd.ExcelWriter(filepath, mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 返回处理结果信息
        return {
            "message": f"批量处理完成，执行了 {len(operations)} 个操作",
            "original_row_count": original_row_count,
            "final_row_count": len(df)
        }
        
    except ImportError:
        raise DataError("批量处理需要安装pandas库：pip install pandas")
    except Exception as e:
        logger.error(f"批量处理数据时出错: {e}")
        raise DataError(str(e))

def convert_format(
    input_filepath: str,
    output_filepath: str,
    output_format: str,
    sheet_name: str = None,
    encoding: str = 'utf-8'
) -> dict:
    """将Excel文件转换为其他格式。
    
    Args:
        input_filepath: 输入Excel文件路径
        output_filepath: 输出文件路径
        output_format: 输出格式 ('csv', 'json', 'html', 'parquet', 'feather')
        sheet_name: 要转换的工作表名称，默认为None（转换所有工作表）
        encoding: 输出文件编码
        
    Returns:
        包含转换结果信息的字典
    """
    try:
        import pandas as pd
        import os
        
        # 验证输出格式
        supported_formats = ['csv', 'json', 'html', 'parquet', 'feather']
        if output_format.lower() not in supported_formats:
            raise DataError(f"不支持的输出格式: {output_format}。支持的格式: {', '.join(supported_formats)}")
        
        # 读取Excel文件
        excel_file = pd.ExcelFile(input_filepath)
        
        # 确定要处理的工作表
        if sheet_name is not None:
            if sheet_name not in excel_file.sheet_names:
                raise DataError(f"工作表 '{sheet_name}' 不存在。可用工作表: {', '.join(excel_file.sheet_names)}")
            sheets_to_process = [sheet_name]
        else:
            sheets_to_process = excel_file.sheet_names
        
        # 根据输出格式确定文件扩展名
        extension_map = {
            'csv': '.csv',
            'json': '.json',
            'html': '.html',
            'parquet': '.parquet',
            'feather': '.feather'
        }
        ext = extension_map[output_format.lower()]
        
        # 处理单个或多个工作表
        if len(sheets_to_process) == 1:
            # 单个工作表处理
            sheet = sheets_to_process[0]
            df = pd.read_excel(input_filepath, sheet_name=sheet)
            
            # 应用转换
            if output_format.lower() == 'csv':
                df.to_csv(output_filepath, index=False, encoding=encoding)
            elif output_format.lower() == 'json':
                df.to_json(output_filepath, orient='records', force_ascii=False)
            elif output_format.lower() == 'html':
                df.to_html(output_filepath, index=False, encoding=encoding)
            elif output_format.lower() == 'parquet':
                df.to_parquet(output_filepath, index=False)
            elif output_format.lower() == 'feather':
                df.to_feather(output_filepath)
            
            return {
                "message": f"已将 '{sheet}' 工作表转换为 {output_format.upper()} 格式",
                "output_file": output_filepath
            }
        else:
            # 多个工作表处理
            output_files = []
            
            # 创建输出目录（如果不存在）
            output_dir = os.path.dirname(output_filepath)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 处理文件名
            base_name = os.path.basename(output_filepath)
            if '.' in base_name:
                base_name = base_name[:base_name.rfind('.')]
            
            # 为每个工作表创建对应的输出文件
            for sheet in sheets_to_process:
                sheet_filename = f"{base_name}_{sheet}{ext}"
                sheet_filepath = os.path.join(os.path.dirname(output_filepath), sheet_filename)
                
                df = pd.read_excel(input_filepath, sheet_name=sheet)
                
                # 应用转换
                if output_format.lower() == 'csv':
                    df.to_csv(sheet_filepath, index=False, encoding=encoding)
                elif output_format.lower() == 'json':
                    df.to_json(sheet_filepath, orient='records', force_ascii=False)
                elif output_format.lower() == 'html':
                    df.to_html(sheet_filepath, index=False, encoding=encoding)
                elif output_format.lower() == 'parquet':
                    df.to_parquet(sheet_filepath, index=False)
                elif output_format.lower() == 'feather':
                    df.to_feather(sheet_filepath)
                
                output_files.append(sheet_filepath)
            
            return {
                "message": f"已将 {len(sheets_to_process)} 个工作表转换为 {output_format.upper()} 格式",
                "output_files": output_files
            }
    
    except ImportError as e:
        missing_lib = str(e).split("'")[1] if "'" in str(e) else "required library"
        raise DataError(f"转换格式需要安装 {missing_lib} 库")
    except Exception as e:
        logger.error(f"转换格式时出错: {e}")
        raise DataError(str(e))
